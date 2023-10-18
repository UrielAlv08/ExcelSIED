from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

archivo_entrada = 'cartera-creditos.xlsx'
entrada_workbook = load_workbook(archivo_entrada, data_only=True)  # Cargar solo los valores, no las fórmulas
entrada_sheet = entrada_workbook.active

# Número de líneas en cada archivo más pequeño
lineas_por_archivo = 5527

# Contador para rastrear el número de líneas procesadas
contador_lineas = 0

# Contador para nombrar los archivos de salida
numero_de_archivo = 1

# Crear un nuevo archivo Excel para el primer conjunto de líneas
archivo_salida = Workbook()
archivo_salida_sheet = archivo_salida.active

# Crear un estilo personalizado para números con 2 decimales
number_style = NamedStyle(name="number_style")
number_style.number_format = '0.00'

for fila in entrada_sheet.iter_rows(values_only=True):
    # Agregar la fila al archivo de salida
    archivo_salida_sheet.append(fila)

    # Aplicar el estilo a todas las celdas que contengan números decimales
    for row in archivo_salida_sheet.iter_rows(min_row=archivo_salida_sheet.max_row):
        for cell in row:
            if isinstance(cell.value, float) and cell.value.is_integer() == False:
                cell.style = number_style

    # Aumentar el contador de líneas
    contador_lineas += 1

    # Si se alcanza el número máximo de líneas por archivo, guardar y crear un nuevo archivo
    if contador_lineas >= lineas_por_archivo:
        nombre_archivo_salida = f'segmento_{numero_de_archivo}.xlsx'
        archivo_salida.save(nombre_archivo_salida)
        archivo_salida.close()

        # Reiniciar el contador de líneas y crear un nuevo archivo
        contador_lineas = 0
        numero_de_archivo += 1
        archivo_salida = Workbook()
        archivo_salida_sheet = archivo_salida.active

# Guardar el último archivo si quedan líneas por procesar
if contador_lineas > 0:
    nombre_archivo_salida = f'segmento_{numero_de_archivo}.xlsx'
    archivo_salida.save(nombre_archivo_salida)
    archivo_salida.close()

print("Ya acabo :D")
entrada_workbook.close()
